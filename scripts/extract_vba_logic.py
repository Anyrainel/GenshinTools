"""Extract character/weapon logic from the Genshin Excel damage calculator.

Characters: Parses VBA character modules, outputs per-character summary focusing
on what's UNIQUE (inline calculations, special mechanics, non-standard formulas).

Weapons: Reads weapon xlsx files, extracts passive skill descriptions.

Usage:
  python extract_vba_logic.py                   → output all to docs/formulas/extracted/
  python extract_vba_logic.py --char 胡桃        → single character to stdout
  python extract_vba_logic.py --weapon 护摩之杖   → single weapon to stdout
  python extract_vba_logic.py --list             → list all characters
  python extract_vba_logic.py --raw 胡桃         → raw VBA block for a character
"""

import re
import sys
from pathlib import Path

import openpyxl
from oletools.olevba import VBA_Parser

XLSM_PATH = Path(__file__).parent / ".." / "docs" / "formulas" / "原神伤害计算(1).xlsm"
WEAPON_DIR = Path(__file__).parent / ".." / "docs" / "formulas" / "武器"
OUTPUT_DIR = Path(__file__).parent / ".." / "docs" / "formulas" / "extracted"

# ── VBA variable → semantic name mapping ──

STAT_MAP = {
    "gongji": "atk",
    "shengming": "hp",
    "fangyu": "def",
    "jingtong": "em",
    "jichugongjili": "baseAtk",
    "chongneng": "er",
}

ELEMENT_MAP = {
    "wushang": "physical",
    "huoshang": "pyro",
    "shuishang": "hydro",
    "fengshang": "anemo",
    "leishang": "electro",
    "caoshang": "dendro",
    "bingshang": "cryo",
    "yanshang": "geo",
}

ABILITY_ZONE_MAP = {
    "pugongbeilvqu": "normal",
    "zhongjibeilvqu": "charged",
    "xialuobeilvqu": "plunge",
    "zhanjibeilvqu": "skill",
    "baofabeilvqu": "burst",
    "pugongjichuqu": "normal",
    "zhongjijichuqu": "charged",
    "xialuojichuqu": "plunge",
    "zhanjijichuqu": "skill",
    "baofajichuqu": "burst",
}

MODULE_ELEMENT = {
    "A火": "pyro",
    "B水": "hydro",
    "C风": "anemo",
    "D雷": "electro",
    "E草": "dendro",
    "F冰": "cryo",
    "G岩": "geo",
}

# 手动输入 sheet special parameter cells → readable names
# Format: label row has description, value row = label row + 1
SPECIAL_PARAM_CELLS: dict[str, str] = {}  # populated at runtime from xlsx

# Fallback hardcoded map (used if xlsx can't be read)
_HARDCODED_PARAMS = {
    "L35": "额外基础生命值",
    "M35": "额外基础攻击力",
    "N35": "额外基础防御力",
    "O35": "额外基础元素精通",
    "L38": "玛薇卡战意值",
    "M38": "阿蕾奇诺生命之契",
    "L41": "芙宁娜0命气氛值",
    "M41": "芙宁娜1命气氛值",
    "N41": "芙宁娜2命气氛值",
    "O41": "芙宁娜2命溢出气氛值",
    "P41": "那维莱特当前生命值",
    "Q41": "水主充盈消耗生命值",
    "L44": "流浪者当前空居力",
    "M44": "伊法队伍夜魂总值",
    "N44": "伊法2命队伍夜魂总值",
    "L47": "克洛琳德生命之契",
    "M47": "雷电将军愿力",
    "N47": "伊安珊夜魂值",
    "L53": "丝柯克蛇之狡谋",
    "M53": "茜特菈莉秘律之数",
    "N53": "申鹤霜霄诀",
    "O53": "优菈能量层数",
    "L56": "兹白时隙浮光",
}


def load_special_params():
    """Load special parameter names from the 手动输入 sheet."""
    global SPECIAL_PARAM_CELLS
    try:
        wb = openpyxl.load_workbook(str(XLSM_PATH), data_only=True, read_only=True)
        ws = wb["手动输入"]
        for row in range(34, 60):
            for col in range(12, 18):  # L=12 through Q=17
                cell = ws.cell(row=row, column=col)
                if cell.value is not None and isinstance(cell.value, str) and len(cell.value) > 1:
                    # Label is here, value is in row+1
                    val_coord = ws.cell(row=row + 1, column=col).coordinate
                    SPECIAL_PARAM_CELLS[val_coord] = cell.value
        wb.close()
    except Exception:
        pass

    if not SPECIAL_PARAM_CELLS:
        SPECIAL_PARAM_CELLS.update(_HARDCODED_PARAMS)


# VBA → English translation table (longest first to avoid partial matches)
VBA_TRANSLATIONS = [
    ("pugongbaojiqiwang", "normalCritExp"),
    ("zhongjibaojiqiwang", "chargedCritExp"),
    ("xialuobaojiqiwang", "plungeCritExp"),
    ("zhanjibaojiqiwang", "skillCritExp"),
    ("baofabaojiqiwang", "burstCritExp"),
    ("fanyingbaojiqiwang", "reactionCritExp"),
    ("pugongbaojiqu", "normalCritMult"),
    ("zhongjibaojiqu", "chargedCritMult"),
    ("xialuobaojiqu", "plungeCritMult"),
    ("zhanjibaojiqu", "skillCritMult"),
    ("baofabaojiqu", "burstCritMult"),
    ("fanyingbaojiqu", "reactionCritMult"),
    ("pugongbeilvqu", "normalMultZone"),
    ("zhongjibeilvqu", "chargedMultZone"),
    ("xialuobeilvqu", "plungeMultZone"),
    ("zhanjibeilvqu", "skillMultZone"),
    ("baofabeilvqu", "burstMultZone"),
    ("zhiliaobeilvqu", "healMultZone"),
    ("hudunbeilvqu", "shieldMultZone"),
    ("pugongjichuqu", "normalBaseZone"),
    ("zhongjijichuqu", "chargedBaseZone"),
    ("xialuojichuqu", "plungeBaseZone"),
    ("zhanjijichuqu", "skillBaseZone"),
    ("baofajichuqu", "burstBaseZone"),
    ("fanyingjichuqu", "reactionBaseZone"),
    ("pugongzengshang", "normalDmgBonus"),
    ("zhongjizengshang", "chargedDmgBonus"),
    ("xialuozengshang", "plungeDmgBonus"),
    ("zhanjizengshang", "skillDmgBonus"),
    ("baofazengshang", "burstDmgBonus"),
    ("quanzengshang", "allDmgBonus"),
    ("dagongjibeilv", "talentMult"),
    ("dashengmingbeilv", "talentMultHP"),
    ("dafangyubeilv", "talentMultDEF"),
    ("dajingtongbeilv", "talentMultEM"),
    ("jichugongjili", "baseAtk"),
    ("zhiliaojiacheng", "healBonus"),
    ("shouzhiliaojiacheng", "incHealBonus"),
    ("hudunqiangxiao", "shieldBonus"),
    ("baojilv", "critRate"),
    ("baojishanghai", "critDmg"),
    ("baojiqiwang", "critExp"),
    ("baojiqu", "critMult"),
    ("fangyuqu", "defMult"),
    ("dengjixishu", "levelCoeff"),
    ("shuihuozhengfa", "vaporize_hydro"),
    ("huoshuizhengfa", "vaporize_pyro"),
    ("huobingronghua", "melt_pyro"),
    ("binghuoronghua", "melt_cryo"),
    ("manjihua", "spread"),
    ("chaojihua", "aggravate"),
    ("wukang", "physicalRes"),
    ("huokang", "pyroRes"),
    ("shuikang", "hydroRes"),
    ("fengkang", "anemoRes"),
    ("leikang", "electroRes"),
    ("caokang", "dendroRes"),
    ("bingkang", "cryoRes"),
    ("yankang", "geoRes"),
]


def translate_vba_expr(expr: str) -> str:
    """Translate VBA variable names to English semantic names."""
    result = expr

    # Replace Worksheets("手动输入").Range("XX").Value → 【参数名】
    for cell_ref, param_name in SPECIAL_PARAM_CELLS.items():
        pattern = rf'Worksheets\("手动输入"\)\.Range\("{cell_ref}"\)\.Value'
        result = re.sub(pattern, f"【{param_name}】", result)
    # Catch any remaining Worksheets references not in our map
    result = re.sub(
        r'Worksheets\("手动输入"\)\.Range\("([^"]+)"\)\.Value',
        r"【手动输入!\1】",
        result,
    )

    for vba, eng in VBA_TRANSLATIONS:
        result = re.sub(rf"\b{vba}\b", eng, result)
    for vba, eng in STAT_MAP.items():
        result = re.sub(rf"\b{vba}\b", eng, result)
    for vba, eng in ELEMENT_MAP.items():
        result = re.sub(rf"\b{eng}shang\b", f"{eng}Dmg", result)
        result = result.replace(vba, f"{eng}Dmg")
    return result


# ──────────────────── Character extraction ────────────────────


def extract_all_vba_modules() -> dict[str, str]:
    vba = VBA_Parser(str(XLSM_PATH))
    modules = {}
    for _fn, _sp, vba_filename, vba_code in vba.extract_macros():
        if "角色模块" in vba_filename:
            modules[vba_filename] = vba_code
    vba.close()
    return modules


def split_character_blocks(modules: dict[str, str]) -> dict[str, dict]:
    characters = {}
    for mod_name, code in modules.items():
        element = "unknown"
        for prefix, elem in MODULE_ELEMENT.items():
            if prefix in mod_name:
                element = elem
                break

        lines = code.split("\n")
        char_starts = []
        for i, line in enumerate(lines):
            m = re.match(r'^(\s{12,20})Case Is = "([^"]+)"', line)
            if m and len(m.group(1)) <= 20:
                lookahead = "\n".join(lines[i : i + 15])
                if (
                    "Select Case .Cells(hang" in lookahead
                    or ".Cells(hang, 4)" in lookahead
                    or ".Cells(hang, 5)" in lookahead
                    or "Worksheets(" in lookahead
                ):
                    char_starts.append((i, m.group(2)))

        for idx, (start_line, char_name) in enumerate(char_starts):
            if idx + 1 < len(char_starts):
                end_line = char_starts[idx + 1][0]
            else:
                end_line = len(lines)
                for j in range(start_line + 1, len(lines)):
                    if re.match(r"^End Function", lines[j]):
                        end_line = j
                        break
            block = "\n".join(lines[start_line:end_line])
            characters[char_name] = {
                "block": block,
                "module": mod_name,
                "element": element,
            }
    return characters


def detect_skill_info(formula: str) -> dict:
    info = {}
    if re.search(r"\bshengming\b\s*\*", formula):
        info["stat"] = "hp"
    elif re.search(r"\bfangyu\b\s*\*", formula):
        info["stat"] = "def"
    elif re.search(r"\bjingtong\b\s*\*", formula):
        info["stat"] = "em"
    else:
        info["stat"] = "atk"
    for vba, elem in ELEMENT_MAP.items():
        if vba in formula:
            info["element"] = elem
            break
    for vba, ability in ABILITY_ZONE_MAP.items():
        if vba in formula:
            info["ability"] = ability
            break
    return info


def is_standard_formula(formula: str) -> bool:
    return bool(
        re.match(
            r"\(gongji \* \(dagongjibeilv \* \w+beilvqu\) \+ \w+jichuqu\) \* ",
            formula.strip(),
        )
    )


def analyze_character_md(char_name: str, data: dict) -> str:
    """Produce compact markdown analysis of a character."""
    block = data["block"]
    element = data["element"]

    out = []
    out.append(f"### {char_name}")
    out.append(f"`{element}` · `{data['module']}`")

    # ── Skills ──
    skill_groups: dict[str, list[str]] = {}
    for m in re.finditer(
        r"Case Is = (.+?)$\s*((?:[ \t]+\.Cells\(hang[^\n]*\n)+)",
        block,
        re.MULTILINE,
    ):
        skill_names = re.findall(r'"([^"]+)"', m.group(1))
        formula_block = m.group(2)
        col4_line = ""
        for fl in formula_block.split("\n"):
            if ".Cells(hang, 4)" in fl:
                col4_line = fl.split("=", 1)[1].strip() if "=" in fl else ""
                break
        if not col4_line:
            continue

        info = detect_skill_info(col4_line)
        stat = info.get("stat", "atk")
        elem = info.get("element", element)
        ability = info.get("ability", "?")
        reactions = set()
        for fl in formula_block.split("\n"):
            rm = re.search(r"\.Cells\(hang,\s*(\d+)\)", fl)
            if rm:
                col = int(rm.group(1))
                if col == 7:
                    reactions.add("vap_p")
                elif col == 10:
                    reactions.add("melt_p")
                elif col == 13:
                    reactions.add("vap_h")
                elif col == 16:
                    reactions.add("melt_c")
                elif col == 19:
                    reactions.add("aggr")

        key_parts = [stat, elem, ability]
        if reactions:
            key_parts.append(",".join(sorted(reactions)))
        if not is_standard_formula(col4_line):
            key_parts.append("★")
        key = "|".join(key_parts)
        if key not in skill_groups:
            skill_groups[key] = []
        skill_groups[key].extend(skill_names)

    out.append("")
    out.append("| Skills | Formula |")
    out.append("|--------|---------|")
    for key, names in skill_groups.items():
        parts = key.split("|")
        stat, elem, ability = parts[0], parts[1], parts[2]
        has_react = len(parts) > 3 and parts[3] not in ("★", "")
        modified = "★" in key

        if len(names) > 3:
            name_str = f"{names[0]}, ... ({len(names)})"
        else:
            name_str = ", ".join(names)

        formula_parts = []
        if stat != "atk":
            formula_parts.append(stat)
        formula_parts.append(elem)
        formula_parts.append(ability)
        formula_str = f"`dmg({', '.join(formula_parts)})`"
        if has_react:
            react_str = parts[3]
            formula_str += f" +{react_str}"
        if modified:
            formula_str += " ★"

        out.append(f"| {name_str} | {formula_str} |")

    # ── Special Logic ──
    dim_vars = re.findall(r"Dim\s+(\w+)", block)
    has_special = dim_vars or re.search(r"Worksheets\(", block) or re.search(r"If\s+.+Then", block)

    # Detect which special params this character reads
    used_params = []
    for cell_ref, param_name in SPECIAL_PARAM_CELLS.items():
        if f'"{cell_ref}"' in block:
            used_params.append((cell_ref, param_name))

    if used_params:
        out.append("")
        out.append("**Params:** " + ", ".join(f"`{name}` ({ref})" for ref, name in used_params))

    if has_special:
        out.append("")
        out.append("**Special Logic:**")
        out.append("```")
        dim_set = set(dim_vars)
        for raw_line in block.split("\n"):
            s = raw_line.strip()
            if not s or s.startswith("'"):
                continue
            is_relevant = (
                s.startswith("Dim ")
                or any(s.startswith(kw) for kw in ("If ", "ElseIf ", "Else", "End If"))
                or ("Worksheets(" in s and ".Value" in s)
                or (
                    ".Cells(" in s
                    and ".Value" in s
                    and "hang" not in s.split(".Cells(")[1].split(")")[0]
                )
                or (
                    "=" in s
                    and not s.startswith(".")
                    and not s.startswith("Case")
                    and not s.startswith("Select")
                    and any(v in s for v in dim_set)
                )
            )
            if is_relevant:
                out.append(translate_vba_expr(s))
        out.append("```")

    # ── Modified Formulas ──
    modified_skills = []
    for m in re.finditer(
        r"Case Is = (.+?)$\s*((?:[ \t]+\.Cells\(hang[^\n]*\n)+)",
        block,
        re.MULTILINE,
    ):
        skill_names = re.findall(r'"([^"]+)"', m.group(1))
        formula_block = m.group(2)
        col4_line = ""
        for fl in formula_block.split("\n"):
            if ".Cells(hang, 4)" in fl:
                col4_line = fl.split("=", 1)[1].strip() if "=" in fl else ""
                break
        if col4_line and not is_standard_formula(col4_line):
            translated = translate_vba_expr(col4_line)
            modified_skills.append((skill_names, translated))

    if modified_skills:
        out.append("")
        out.append("**Modified Formulas:**")
        out.append("```")
        for names, formula in modified_skills:
            label = names[0] if len(names) == 1 else f"{names[0]} (+{len(names) - 1})"
            if len(formula) > 100:
                formula = formula[:97] + "..."
            out.append(f"[{label}]")
            out.append(f"  {formula}")
        out.append("```")

    out.append("")
    return "\n".join(out)


# ──────────────────── Weapon extraction ────────────────────

WEAPON_TYPES = {
    "单手剑": "Sword",
    "双手剑": "Claymore",
    "弓": "Bow",
    "法器": "Catalyst",
    "长柄武器": "Polearm",
}

SUBSTAT_MAP = {
    "大攻击": "ATK%",
    "暴击率": "CritRate",
    "暴击伤害": "CritDMG",
    "精通": "EM",
    "充能": "ER",
    "大生命": "HP%",
    "大防御": "DEF%",
    "物伤": "PhysDMG%",
}


def extract_weapon(xlsx_path: Path) -> dict | None:
    """Extract weapon data from its xlsx file."""
    try:
        wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
        ws = wb["基础面板"]

        name = xlsx_path.stem
        passive_name = ws["G1"].value or ""
        passive_desc = ws["G2"].value or ""
        substat_raw = ws["D1"].value or ""
        base_atk_90 = ws["B10"].value  # Lv90

        # Clean passive name
        if passive_name.startswith("武器技能 - "):
            passive_name = passive_name[len("武器技能 - ") :]

        # Map substat
        substat = SUBSTAT_MAP.get(substat_raw, substat_raw)

        return {
            "name": name,
            "passive_name": passive_name,
            "passive_desc": passive_desc,
            "substat": substat,
            "base_atk_90": base_atk_90,
        }
    except Exception as e:
        return {"name": xlsx_path.stem, "error": str(e)}


def extract_all_weapons() -> dict[str, list[dict]]:
    """Extract all weapons grouped by type."""
    result = {}
    for type_dir in sorted(WEAPON_DIR.iterdir()):
        if not type_dir.is_dir():
            continue
        type_name = type_dir.name
        weapons = []
        for xlsx_path in sorted(type_dir.glob("*.xlsx")):
            data = extract_weapon(xlsx_path)
            if data:
                weapons.append(data)
        result[type_name] = weapons
    return result


def format_weapon_md(weapon: dict) -> str:
    """Format a single weapon as markdown."""
    if "error" in weapon:
        return f"#### {weapon['name']}\n_Error: {weapon['error']}_\n"

    out = []
    out.append(f"#### {weapon['name']}")

    meta = []
    if weapon["base_atk_90"]:
        meta.append(f"BaseATK={weapon['base_atk_90']}")
    if weapon["substat"]:
        meta.append(f"Sub={weapon['substat']}")
    if weapon["passive_name"]:
        meta.append(f"「{weapon['passive_name']}」")
    out.append(" · ".join(meta))

    if weapon["passive_desc"]:
        # Parse refinement values from the description
        # Format: "提升20%/25%/30%/35%/40%" → keep as-is, it's compact enough
        desc = weapon["passive_desc"]
        out.append("")
        out.append(f"> {desc}")

    out.append("")
    return "\n".join(out)


# ──────────────────── Main ────────────────────


def main():
    args = sys.argv[1:]
    load_special_params()

    if "--weapon" in args:
        idx = args.index("--weapon")
        weapon_name = args[idx + 1]
        # Search all weapon dirs
        for type_dir in WEAPON_DIR.iterdir():
            if not type_dir.is_dir():
                continue
            path = type_dir / f"{weapon_name}.xlsx"
            if path.exists():
                data = extract_weapon(path)
                if data:
                    print(format_weapon_md(data))
                return
        print(f"Weapon '{weapon_name}' not found")
        return

    modules = extract_all_vba_modules()
    characters = split_character_blocks(modules)

    if "--list" in args:
        for name, data in sorted(characters.items()):
            has_special = bool(re.findall(r"Dim\s+\w+", data["block"])) or bool(
                re.findall(r"Worksheets\(", data["block"])
            )
            tag = " ★" if has_special else ""
            print(f"  {name} ({data['element']}){tag}")
        print(f"\nTotal: {len(characters)} characters (★ = has special logic)")
        return

    if "--raw" in args:
        idx = args.index("--raw")
        char_name = args[idx + 1]
        if char_name in characters:
            print(characters[char_name]["block"])
        else:
            print(f"Character '{char_name}' not found")
        return

    if "--char" in args:
        idx = args.index("--char")
        char_name = args[idx + 1]
        if char_name in characters:
            print(analyze_character_md(char_name, characters[char_name]))
        else:
            print(f"Character '{char_name}' not found")
            print(f"Available: {', '.join(sorted(characters.keys()))}")
        return

    # ── Default: output all to docs/formulas/extracted/ ──
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    ELEMENT_NAMES = {
        "pyro": "火元素",
        "hydro": "水元素",
        "anemo": "风元素",
        "electro": "雷元素",
        "dendro": "草元素",
        "cryo": "冰元素",
        "geo": "岩元素",
    }

    # ── Characters ──
    by_element: dict[str, list[tuple[str, dict]]] = {}
    for name, data in sorted(characters.items()):
        elem = data["element"]
        if elem not in by_element:
            by_element[elem] = []
        by_element[elem].append((name, data))

    for elem, chars in sorted(by_element.items()):
        cn_name = ELEMENT_NAMES.get(elem, elem)
        filepath = OUTPUT_DIR / f"角色-{cn_name}.md"
        lines = [
            f"# {cn_name} · 角色伤害逻辑",
            "",
            "> Extracted from `docs/formulas/原神伤害计算(1).xlsm` VBA",
            "> Generated by `scripts/extract_vba_logic.py`",
            "",
            "**Standard formula:** `(stat × talentMult × multZone + baseZone) × (elemDmg% + allDmg% + abilityDmg%) × crit × def × res`",
            "",
            "- `dmg(element, ability)` = standard ATK-scaled formula",
            "- `dmg(hp, element, ability)` = HP-scaled",
            "- `dmg(def, element, ability)` = DEF-scaled",
            "- ★ = formula modified inline (see Modified Formulas section)",
            "- `【参数名】` = special parameter read from 手动输入 sheet (user input)",
            "",
            "---",
            "",
        ]
        for name, data in chars:
            lines.append(analyze_character_md(name, data))
        filepath.write_text("\n".join(lines), encoding="utf-8")
        print(f"Wrote {filepath} ({len(chars)} characters)")

    # ── Special Params Reference ──
    params_path = OUTPUT_DIR / "参数表.md"
    params_lines = [
        "# 手动输入 · 专用参数表",
        "",
        '> 角色 VBA 中通过 `Worksheets("手动输入").Range("XX").Value` 读取的特殊参数。',
        "> 这些需要用户手动填入，计算器不会自动计算。",
        "",
        "| Cell | 参数名 | 默认值 |",
        "|------|--------|--------|",
    ]
    try:
        wb = openpyxl.load_workbook(str(XLSM_PATH), data_only=True, read_only=True)
        ws = wb["手动输入"]
        for cell_ref, param_name in sorted(SPECIAL_PARAM_CELLS.items()):
            # Read current value
            from openpyxl.utils import coordinate_to_tuple

            row, col = coordinate_to_tuple(cell_ref)
            val = ws.cell(row=row, column=col).value
            val_str = str(val) if val is not None else ""
            params_lines.append(f"| {cell_ref} | {param_name} | {val_str} |")
        wb.close()
    except Exception:
        for cell_ref, param_name in sorted(SPECIAL_PARAM_CELLS.items()):
            params_lines.append(f"| {cell_ref} | {param_name} | |")

    # Also list buff input rows
    params_lines.extend(
        [
            "",
            "## Buff 输入区域 (C-K 列, 9 个槽位)",
            "",
            "以下 buff 类型需要用户手动填入对应数值。偶数行=值，奇数行=备注。",
            "",
            "| Row | Buff 类型 |",
            "|-----|----------|",
        ]
    )
    try:
        wb = openpyxl.load_workbook(str(XLSM_PATH), data_only=True, read_only=True)
        ws = wb["手动输入"]
        for row_num in range(34, 183, 2):
            label = ws.cell(row=row_num, column=2).value
            if label:
                params_lines.append(f"| {row_num} | {label} |")
        wb.close()
    except Exception:
        pass

    params_path.write_text("\n".join(params_lines), encoding="utf-8")
    print(f"Wrote {params_path}")

    # ── Weapons ──
    all_weapons = extract_all_weapons()
    filepath = OUTPUT_DIR / "武器.md"
    lines = [
        "# 武器被动技能",
        "",
        "> Extracted from `docs/formulas/武器/**/*.xlsx`",
        "> Generated by `scripts/extract_vba_logic.py`",
        "",
        "**Note:** 在这个Excel计算器中，武器特效不在VBA中计算——需要用户手动在「手动输入」",
        "sheet的buff栏填入对应数值。此文件仅列出武器被动描述文本，用于交叉验证项目实现。",
        "",
        "Refinement values are listed as R1/R2/R3/R4/R5.",
        "",
        "---",
        "",
    ]
    total_weapons = 0
    for type_cn, weapons in sorted(all_weapons.items()):
        type_en = WEAPON_TYPES.get(type_cn, type_cn)
        lines.append(f"## {type_cn} ({type_en})")
        lines.append("")
        for w in weapons:
            lines.append(format_weapon_md(w))
        total_weapons += len(weapons)

    filepath.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {filepath} ({total_weapons} weapons)")

    print(f"\nTotal: {len(characters)} characters, {total_weapons} weapons")


if __name__ == "__main__":
    main()
