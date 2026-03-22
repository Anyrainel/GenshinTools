"""Inspect Excel files from the Genshin damage calculator spreadsheets.

Usage:
  python inspect_excel.py overview <file>        - Sheet names, dimensions, named ranges
  python inspect_excel.py dump <file> [sheet]     - Dump cell values + formulas for a sheet
  python inspect_excel.py formulas <file> [sheet] - Show only cells with formulas
  python inspect_excel.py names <file>            - Show all defined names (named ranges)
  python inspect_excel.py cross-refs <file>       - Show cross-sheet/cross-file references
"""

import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def overview(path: str):
    wb = load_workbook(path, data_only=False)
    print(f"=== {Path(path).name} ===")
    print(f"Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")
    print()

    # Defined names
    if wb.defined_names:
        print("Defined Names:")
        for dn in wb.defined_names.definedName:
            print(f"  {dn.name} = {dn.attr_text}")
        print()

    for name in wb.sheetnames:
        ws = wb[name]
        print(f"--- Sheet: {name} ---")
        print(f"  Dimensions: {ws.dimensions}")
        print(f"  Max row: {ws.max_row}, Max col: {ws.max_column}")

        # Count formulas
        formula_count = 0
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
        print(f"  Formula cells: {formula_count}")

        # Merged cells
        if ws.merged_cells.ranges:
            print(f"  Merged ranges: {len(ws.merged_cells.ranges)}")

        print()
    wb.close()


def dump_sheet(path: str, sheet_name: str | None = None, max_rows: int = 200):
    wb = load_workbook(path, data_only=False)
    wb_data = load_workbook(path, data_only=True)

    if sheet_name is None:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]
    ws_data = wb_data[sheet_name]

    print(f"=== {Path(path).name} / {sheet_name} ===")
    print(f"Dimensions: {ws.dimensions}, Max row: {ws.max_row}, Max col: {ws.max_column}")
    print()

    row_count = 0
    for row_idx, (row_f, row_v) in enumerate(
        zip(ws.iter_rows(), ws_data.iter_rows(), strict=True), start=1
    ):
        if row_count >= max_rows:
            print(f"... truncated at {max_rows} rows ...")
            break

        cells = []
        for cell_f, cell_v in zip(row_f, row_v, strict=True):
            coord = f"{get_column_letter(cell_f.column)}{cell_f.row}"
            val_f = cell_f.value  # formula or literal
            val_v = cell_v.value  # computed value

            if val_f is None and val_v is None:
                continue

            if isinstance(val_f, str) and val_f.startswith("="):
                cells.append(f"  {coord}: {val_f}  (= {val_v})")
            else:
                cells.append(f"  {coord}: {repr(val_f)}")

        if cells:
            row_count += 1
            print(f"Row {row_idx}:")
            for c in cells:
                print(c)

    wb.close()
    wb_data.close()


def show_formulas(path: str, sheet_name: str | None = None):
    wb = load_workbook(path, data_only=False)
    wb_data = load_workbook(path, data_only=True)

    sheets = [sheet_name] if sheet_name else wb.sheetnames

    for sn in sheets:
        ws = wb[sn]
        ws_data = wb_data[sn]
        print(f"=== {Path(path).name} / {sn} ===")

        for row_f, row_v in zip(ws.iter_rows(), ws_data.iter_rows(), strict=True):
            for cell_f, cell_v in zip(row_f, row_v, strict=True):
                if isinstance(cell_f.value, str) and cell_f.value.startswith("="):
                    coord = f"{get_column_letter(cell_f.column)}{cell_f.row}"
                    print(f"  {coord}: {cell_f.value}  (= {cell_v.value})")
        print()

    wb.close()
    wb_data.close()


def show_names(path: str):
    wb = load_workbook(path, data_only=False)
    print(f"=== Defined Names in {Path(path).name} ===")
    if not wb.defined_names:
        print("  (none)")
    else:
        for dn in wb.defined_names.definedName:
            print(f"  {dn.name} = {dn.attr_text}")
    wb.close()


def cross_refs(path: str):
    """Find formulas that reference other sheets or external files."""
    wb = load_workbook(path, data_only=False)

    # Patterns for cross-sheet refs like Sheet!A1 or 'Sheet Name'!A1
    sheet_ref_pat = re.compile(r"(?:'([^']+)'|(\w+))!")
    # Pattern for external file refs like [file.xlsx]Sheet!A1
    ext_ref_pat = re.compile(r"\[([^\]]+)\]")

    print(f"=== Cross-references in {Path(path).name} ===")

    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
                    coord = f"{get_column_letter(cell.column)}{cell.row}"

                    ext_matches = ext_ref_pat.findall(formula)
                    if ext_matches:
                        print(f"  [{sn}!{coord}] EXTERNAL: {formula}")
                        print(f"    -> files: {ext_matches}")

                    sheet_matches = sheet_ref_pat.findall(formula)
                    for quoted, unquoted in sheet_matches:
                        ref_sheet = quoted or unquoted
                        if ref_sheet != sn:
                            print(f"  [{sn}!{coord}] -> sheet '{ref_sheet}': {formula}")
                            break

    wb.close()


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    cmd = sys.argv[1]
    path = sys.argv[2]
    sheet = sys.argv[3] if len(sys.argv) > 3 else None

    match cmd:
        case "overview":
            overview(path)
        case "dump":
            dump_sheet(path, sheet)
        case "formulas":
            show_formulas(path, sheet)
        case "names":
            show_names(path)
        case "cross-refs":
            cross_refs(path)
        case _:
            print(f"Unknown command: {cmd}")
            print(__doc__)
